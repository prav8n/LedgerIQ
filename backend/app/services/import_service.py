"""CSV / XLSX import parsing, column auto-mapping and value coercion.

Pure functions (no DB) so they're easy to unit-test. The route layer turns the
mapped rows into expenses via the categorization + cashback services.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

# Target field -> ordered list of header-name substrings used to auto-map.
FIELD_SYNONYMS: dict[str, list[str]] = {
    "transaction_date": ["transaction date", "txn date", "date", "posted", "value date"],
    "amount": ["amount", "debit", "value", "price", "total", "spent"],
    "merchant": ["merchant", "payee", "vendor", "to", "name", "particulars"],
    "description": ["description", "narration", "details", "remarks", "note"],
    "category": ["category", "type", "tag"],
    "payment_method": ["payment", "method", "mode", "channel"],
}

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
    "%d.%m.%Y", "%d %b %Y", "%d-%b-%Y", "%d %B %Y", "%b %d, %Y",
    "%Y-%m-%d %H:%M:%S", "%d/%m/%y", "%d-%m-%y",
]


def _norm(text: str) -> str:
    return text.strip().lower()


def suggest_mapping(headers: list[str]) -> dict[str, str | None]:
    """Best-guess mapping of each target field to a source header."""
    normalized = {h: _norm(h) for h in headers}
    used: set[str] = set()
    mapping: dict[str, str | None] = {}
    for field, synonyms in FIELD_SYNONYMS.items():
        match: str | None = None
        # Exact-ish first (synonym equals header), then substring.
        for syn in synonyms:
            for header, norm in normalized.items():
                if header in used:
                    continue
                if norm == syn:
                    match = header
                    break
            if match:
                break
        if match is None:
            for syn in synonyms:
                for header, norm in normalized.items():
                    if header in used:
                        continue
                    if syn in norm:
                        match = header
                        break
                if match:
                    break
        if match:
            used.add(match)
        mapping[field] = match
    return mapping


def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Return ``(headers, rows)`` from a CSV or XLSX upload.

    Rows are dicts keyed by header. Raises ``ValueError`` on unsupported types
    or empty files.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv(content)
    # Fall back to CSV sniffing.
    return _parse_csv(content)


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any(cell.strip() for cell in r)]
    if not rows:
        raise ValueError("The file is empty")
    headers = [h.strip() for h in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        out.append({headers[i]: (raw[i] if i < len(raw) else "") for i in range(len(headers))})
    return headers, out


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook  # imported lazily

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("The workbook has no active sheet")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("The file is empty") from exc
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    out: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            val = raw[i] if i < len(raw) else None
            row[h] = "" if val is None else str(val)
        out.append(row)
    wb.close()
    return headers, out


def coerce_date(value: str) -> date | None:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    # ISO with time / timezone.
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def coerce_amount(value: str) -> Decimal | None:
    value = (value or "").strip()
    if not value:
        return None
    cleaned = (
        value.replace("₹", "")  # ₹
        .replace("INR", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .replace(",", "")
        .replace(" ", "")
    )
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    try:
        amount = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    amount = abs(amount)
    if negative:
        amount = -amount
    return amount
